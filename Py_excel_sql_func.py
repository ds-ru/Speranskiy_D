import sqlite3
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def create_table(connection, cursor):
    connection = connection
    cursor = cursor
    create_main_table = ("CREATE TABLE IF NOT EXISTS maint "
                         "(id INTEGER PRIMARY KEY AUTOINCREMENT, "
                         "main_id VARCHAR,"
                         "region VARCHAR, "
                         "full_name VARCHAR, "
                         "quantity INT, "
                         "price_nds FLOAT, "
                         "brand VARCHAR, "
                         "weight FLOAT)")
    cursor.execute(create_main_table)
    connection.commit()
    print("УСПЕХ")


def read_date(connection, cursor):
    print("Выбрано чтение данных")
    connection = connection
    cursor = cursor
    sheet_name = 'Данные'
    columns = ["main_id", "region", "full_name", "quantity", "price_nds", "brand", "weight"]
    data = pd.read_excel("x5.xlsx", sheet_name, usecols=columns)
    for row in data.itertuples():
        cursor.execute("INSERT INTO maint (main_id, region, full_name, quantity, price_nds, brand, weight)"
                       "VALUES (?, ?, ?, ?, ?, ?, ?)",
                       (row.main_id, row.region, row.full_name, row.quantity, row.price_nds, row.brand, row.weight))
    connection.commit()
    print("УСПЕХ")


def use_code(connection, cursor):
    print("Выбрано использование кода")
    filename = 'D:\\PythonProjectSQL\\test.xlsx'
    sheet = 'Лист1'
    wb = load_workbook(filename)
    ws = wb[sheet]
    ws.delete_rows(1, ws.max_row)
    while True:
        connection = connection
        cursor = cursor
        code = input("Введите код: ")
        if code == "exit":
            break
        cursor.execute(str(code))
        rows = cursor.fetchall()
        print("УСПЕХ")
        for row in rows:
            print(row)
        data_frame = pd.DataFrame(rows, columns=[description[0] for description in cursor.description])
        for r in dataframe_to_rows(data_frame, index=False, header=True):
            ws.append(r)
        wb.save(filename)
        wb.close()


def all_data(connection, cursor):
    print("Выбрано вывод всех данных")
    connection = connection
    cursor = cursor
    cursor.execute("SELECT * FROM maint")
    rows = cursor.fetchall()
    print("УСПЕХ")
    for row in rows:
        print(row)


def test():
    print("УСПЕХ")
    os.system('cls')
    print("УСПЕХ")
